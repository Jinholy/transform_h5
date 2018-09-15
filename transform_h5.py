import h5py as h5
import openpyxl as xl


def read_h5(file_path):
    f = h5.File(file_path, 'r')
    return f


def get_XCO2(h_file, file_name):
    if 'V0246' in file_name:
        return h_file["Data"]["mixingRatio"]["XCO2BiasCorrected"]
    else:
        return h_file["Data"]["mixingRatio"]["XCO2"]


def write_xlsx(x_data, y_data, xco2_data, file_name):
    wb = xl.Workbook()
    work_sheet = wb.active
    work_sheet["A1"] = "x"
    work_sheet["B1"] = "y"
    work_sheet["C1"] = "XCO2"

    for x in range(1, len(x_data)):
        work_sheet.cell(x + 1, 1, value=x_data[x])

    for y in range(1, len(y_data)):
        work_sheet.cell(y + 1, 2, value=y_data[y])

    for xco in range(1, len(xco2_data)):
        work_sheet.cell(xco + 1, 3, value=xco2_data[xco])

    wb.save(file_name)


def get_xlsxname(file_name):
    fl = file_name.split("_")
    date = fl[0][9:18]
    if 'V0226' in file_name:
        return date + "_v2.26" + ".xlsx"
    elif 'V0236' in file_name:
        return date + "_v2.36" + ".xlsx"
    elif 'V0246' in file_name:
        return date + "_v2.46" + ".xlsx"
    elif 'V0250' in file_name:
        return date + "_v2.50" + ".xlsx"
    else:
        return date + "_v2.60" + ".xlsx"


list = list(open("./SWIRL2CO2/list.txt", 'r'))

for path in list:
    f_path = path.replace("\\", "/")[:-1]
    f_name = f_path.split("/")[-1]

    h5_file = read_h5(f_path)

    x_data = (h5_file["Data"]["geolocation"]["longitude"])
    y_data = (h5_file["Data"]["geolocation"]["latitude"])
    XCO2 = get_XCO2(h5_file, f_name)

    write_xlsx(x_data, y_data, XCO2, get_xlsxname(f_name))
